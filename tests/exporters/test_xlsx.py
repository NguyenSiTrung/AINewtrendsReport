"""Tests for the Excel (.xlsx) exporter."""

from __future__ import annotations

from pathlib import Path
from typing import Any

import pytest
from openpyxl import load_workbook
from pydantic import ValidationError

from ainews.schemas.report_output import XlsxOutput

# ── Sample data ──────────────────────────────────────────


def _make_export_data(
    *,
    summaries: list[dict[str, Any]] | None = None,
    trends: list[dict[str, Any]] | None = None,
    executive_summary: str = "Executive summary text.",
    params: dict[str, Any] | None = None,
) -> dict[str, Any]:
    """Build sample export data dict."""
    return {
        "executive_summary": executive_summary,
        "params": params or {"topics": ["AI", "ML"], "timeframe_days": 7},
        "generated_at": "2026-05-08 10:00 UTC",
        "summaries": summaries
        if summaries is not None
        else [
            {
                "cluster_id": "c1",
                "headline": "GPT-5 Released",
                "bullets": ["Faster than GPT-4", "Better reasoning"],
                "why_it_matters": "Major milestone",
                "sources": [
                    "https://openai.com/blog",
                    "https://techcrunch.com/gpt5",
                ],
            },
            {
                "cluster_id": "c2",
                "headline": "EU AI Act Enforced",
                "bullets": ["New compliance requirements"],
                "why_it_matters": "Regulatory impact",
                "sources": ["https://eu-policy.org"],
            },
        ],
        "trends": trends
        if trends is not None
        else [
            {
                "name": "Regulatory Convergence",
                "description": "Global AI regulations are converging.",
                "evidence_cluster_ids": ["c1", "c2"],
            },
        ],
    }


# ── Tests ─────────────────────────────────────────────────


class TestExportXlsx:
    """Tests for export_xlsx function."""

    def test_workbook_has_four_sheets(self, tmp_path: Path) -> None:
        """Exported workbook has exactly 4 sheets."""
        from ainews.exporters.xlsx import export_xlsx

        data = _make_export_data()
        path = export_xlsx(data, "run-001", tmp_path)
        wb = load_workbook(path)
        assert wb.sheetnames == ["Summary", "Stories", "Sources", "Trends"]
        wb.close()

    def test_freeze_panes(self, tmp_path: Path) -> None:
        """All sheets have frozen panes (row 2)."""
        from ainews.exporters.xlsx import export_xlsx

        data = _make_export_data()
        path = export_xlsx(data, "run-001", tmp_path)
        wb = load_workbook(path)
        for name in wb.sheetnames:
            ws = wb[name]
            assert ws.freeze_panes == "A2", f"Freeze panes missing on {name}"
        wb.close()

    def test_header_row_bold(self, tmp_path: Path) -> None:
        """Header rows on data sheets have bold styling."""
        from ainews.exporters.xlsx import export_xlsx

        data = _make_export_data()
        path = export_xlsx(data, "run-001", tmp_path)
        wb = load_workbook(path)
        for name in ["Stories", "Sources", "Trends"]:
            ws = wb[name]
            for cell in ws[1]:
                if cell.value:
                    assert cell.font.bold, f"Header not bold: {name}!{cell.coordinate}"
        wb.close()

    def test_header_row_fill(self, tmp_path: Path) -> None:
        """Header rows on data sheets have a background fill."""
        from ainews.exporters.xlsx import export_xlsx

        data = _make_export_data()
        path = export_xlsx(data, "run-001", tmp_path)
        wb = load_workbook(path)
        for name in ["Stories", "Sources", "Trends"]:
            ws = wb[name]
            first_cell = ws.cell(row=1, column=1)
            # Should have a solid fill
            assert first_cell.fill.fgColor is not None
        wb.close()

    def test_sources_hyperlinks(self, tmp_path: Path) -> None:
        """Sources sheet URL column contains hyperlinks."""
        from ainews.exporters.xlsx import export_xlsx

        data = _make_export_data()
        path = export_xlsx(data, "run-001", tmp_path)
        wb = load_workbook(path)
        ws = wb["Sources"]
        # Check row 2 (first data row) URL column (column A)
        cell = ws.cell(row=2, column=1)
        assert cell.hyperlink is not None, "URL cell should have hyperlink"
        assert cell.hyperlink.target == "https://openai.com/blog"
        wb.close()

    def test_column_width_capped(self, tmp_path: Path) -> None:
        """Auto-sized columns are capped at 80 characters."""
        from ainews.exporters.xlsx import export_xlsx

        data = _make_export_data()
        # Add a story with very long text
        data["summaries"].append(
            {
                "cluster_id": "c3",
                "headline": "X" * 200,
                "bullets": ["A" * 200],
                "why_it_matters": "B" * 200,
                "sources": [],
            }
        )
        path = export_xlsx(data, "run-001", tmp_path)
        wb = load_workbook(path)
        for ws in wb.worksheets:
            for col_dim in ws.column_dimensions.values():
                if col_dim.width:
                    assert col_dim.width <= 82, (  # 80 + small padding
                        f"Column too wide in {ws.title}: {col_dim.width}"
                    )
        wb.close()

    def test_empty_summaries(self, tmp_path: Path) -> None:
        """Excel export handles empty summaries."""
        from ainews.exporters.xlsx import export_xlsx

        data = _make_export_data(summaries=[])
        path = export_xlsx(data, "run-001", tmp_path)
        wb = load_workbook(path)
        ws = wb["Stories"]
        # Should only have header row
        assert ws.max_row == 1
        wb.close()

    def test_empty_trends(self, tmp_path: Path) -> None:
        """Excel export handles empty trends."""
        from ainews.exporters.xlsx import export_xlsx

        data = _make_export_data(trends=[])
        path = export_xlsx(data, "run-001", tmp_path)
        wb = load_workbook(path)
        ws = wb["Trends"]
        assert ws.max_row == 1
        wb.close()

    def test_valid_xlsx_reload(self, tmp_path: Path) -> None:
        """Generated file can be loaded back by openpyxl."""
        from ainews.exporters.xlsx import export_xlsx

        data = _make_export_data()
        path = export_xlsx(data, "run-001", tmp_path)
        wb = load_workbook(path)
        assert len(wb.sheetnames) == 4
        wb.close()

    def test_file_path_is_absolute(self, tmp_path: Path) -> None:
        """Returned file path is absolute."""
        from ainews.exporters.xlsx import export_xlsx

        data = _make_export_data()
        path = export_xlsx(data, "run-001", tmp_path)
        assert path.is_absolute()

    def test_summary_sheet_content(self, tmp_path: Path) -> None:
        """Summary sheet contains metadata and executive summary."""
        from ainews.exporters.xlsx import export_xlsx

        data = _make_export_data()
        path = export_xlsx(data, "run-001", tmp_path)
        wb = load_workbook(path)
        ws = wb["Summary"]
        cell_values = [
            row[0].value
            for row in ws.iter_rows(min_col=2, max_col=2, values_only=False)
        ]
        assert "2026-05-08 10:00 UTC" in cell_values
        assert "AI, ML" in cell_values
        wb.close()

    def test_stories_row_count(self, tmp_path: Path) -> None:
        """Stories sheet has correct number of data rows."""
        from ainews.exporters.xlsx import export_xlsx

        data = _make_export_data()
        path = export_xlsx(data, "run-001", tmp_path)
        wb = load_workbook(path)
        ws = wb["Stories"]
        # Header + 2 summaries = 3 rows
        assert ws.max_row == 3
        wb.close()


# ── XlsxOutput validation tests ─────────────────────────


class TestXlsxOutputValidation:
    """Tests for XlsxOutput Pydantic schema."""

    def test_valid_file(self, tmp_path: Path) -> None:
        """Valid xlsx file passes validation."""
        from ainews.exporters.xlsx import export_xlsx

        data = _make_export_data()
        path = export_xlsx(data, "run-001", tmp_path)
        result = XlsxOutput(file_path=path)
        assert result.file_path == path

    def test_nonexistent_file_rejected(self, tmp_path: Path) -> None:
        """Non-existent file is rejected."""
        fake_path = tmp_path / "nonexistent.xlsx"
        with pytest.raises(ValidationError, match="does not exist"):
            XlsxOutput(file_path=fake_path)

    def test_empty_file_rejected(self, tmp_path: Path) -> None:
        """Empty file is rejected."""
        empty = tmp_path / "empty.xlsx"
        empty.write_bytes(b"")
        with pytest.raises(ValidationError, match="empty"):
            XlsxOutput(file_path=empty)
