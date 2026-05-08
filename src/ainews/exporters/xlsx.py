"""Excel (.xlsx) report exporter.

Builds an openpyxl workbook with four sheets: Summary, Stories,
Sources, and Trends. Formatting constants are extracted to a
module-level ``STYLES`` dict for easy future customization.
"""

from __future__ import annotations

from pathlib import Path
from typing import Any

import structlog
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from ainews.schemas.report_output import XlsxOutput

logger = structlog.get_logger(__name__)


# ── Formatting constants ──────────────────────────────────

STYLES: dict[str, Any] = {
    "header_font": Font(name="Calibri", bold=True, size=11, color="FFFFFF"),
    "header_fill": PatternFill(
        start_color="2F5496", end_color="2F5496", fill_type="solid"
    ),
    "header_alignment": Alignment(horizontal="left", vertical="center"),
    "data_font": Font(name="Calibri", size=11),
    "data_alignment": Alignment(horizontal="left", vertical="top", wrap_text=True),
    "label_font": Font(name="Calibri", bold=True, size=11),
    "max_column_width": 80,
}


# ── Sheet builders ────────────────────────────────────────


def _build_summary_sheet(ws: Worksheet, data: dict[str, Any]) -> None:
    """Build the Summary sheet with metadata and executive summary."""
    ws.title = "Summary"

    meta_rows = [
        ("Report Date", data.get("generated_at", "")),
        ("Topics", ", ".join(data.get("params", {}).get("topics", []))),
        ("Timeframe (days)", data.get("params", {}).get("timeframe_days", "")),
        ("", ""),
        ("Executive Summary", data.get("executive_summary", "")),
    ]

    for row_idx, (label, value) in enumerate(meta_rows, 1):
        label_cell = ws.cell(row=row_idx, column=1, value=label)
        label_cell.font = STYLES["label_font"]
        value_cell = ws.cell(row=row_idx, column=2, value=value)
        value_cell.font = STYLES["data_font"]
        value_cell.alignment = STYLES["data_alignment"]

    # Header row freeze for consistency
    ws.freeze_panes = "A2"

    # Column widths
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 80


def _build_stories_sheet(ws: Worksheet, summaries: list[dict[str, Any]]) -> None:
    """Build the Stories sheet: one row per story."""
    ws.title = "Stories"
    headers = ["Headline", "Key Points", "Why It Matters", "Source Count"]

    _write_header_row(ws, headers)

    for row_idx, summary in enumerate(summaries, 2):
        ws.cell(row=row_idx, column=1, value=summary.get("headline", ""))
        ws.cell(
            row=row_idx,
            column=2,
            value="; ".join(summary.get("bullets", [])),
        )
        ws.cell(row=row_idx, column=3, value=summary.get("why_it_matters", ""))
        ws.cell(
            row=row_idx,
            column=4,
            value=len(summary.get("sources", [])),
        )

        # Style data cells
        for col in range(1, 5):
            cell = ws.cell(row=row_idx, column=col)
            cell.font = STYLES["data_font"]
            cell.alignment = STYLES["data_alignment"]

    _auto_size_columns(ws, max_width=STYLES["max_column_width"])


def _build_sources_sheet(ws: Worksheet, summaries: list[dict[str, Any]]) -> None:
    """Build the Sources sheet: one row per source URL across all stories."""
    ws.title = "Sources"
    headers = ["URL", "Title", "Cluster ID"]

    _write_header_row(ws, headers)

    row_idx = 2
    for summary in summaries:
        cluster_id = summary.get("cluster_id", "")
        headline = summary.get("headline", "")
        for url in summary.get("sources", []):
            url_cell = ws.cell(row=row_idx, column=1, value=url)
            url_cell.hyperlink = url
            url_cell.font = Font(
                name="Calibri", size=11, color="0563C1", underline="single"
            )
            ws.cell(row=row_idx, column=2, value=headline)
            ws.cell(row=row_idx, column=3, value=cluster_id)

            for col in range(2, 4):
                cell = ws.cell(row=row_idx, column=col)
                cell.font = STYLES["data_font"]
                cell.alignment = STYLES["data_alignment"]

            row_idx += 1

    _auto_size_columns(ws, max_width=STYLES["max_column_width"])


def _build_trends_sheet(ws: Worksheet, trends: list[dict[str, Any]]) -> None:
    """Build the Trends sheet: one row per trend."""
    ws.title = "Trends"
    headers = ["Name", "Description", "Evidence Cluster IDs"]

    _write_header_row(ws, headers)

    for row_idx, trend in enumerate(trends, 2):
        ws.cell(row=row_idx, column=1, value=trend.get("name", ""))
        ws.cell(row=row_idx, column=2, value=trend.get("description", ""))
        ws.cell(
            row=row_idx,
            column=3,
            value=", ".join(trend.get("evidence_cluster_ids", [])),
        )

        for col in range(1, 4):
            cell = ws.cell(row=row_idx, column=col)
            cell.font = STYLES["data_font"]
            cell.alignment = STYLES["data_alignment"]

    _auto_size_columns(ws, max_width=STYLES["max_column_width"])


# ── Helpers ───────────────────────────────────────────────


def _write_header_row(ws: Worksheet, headers: list[str]) -> None:
    """Write a styled header row and freeze panes."""
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = STYLES["header_font"]
        cell.fill = STYLES["header_fill"]
        cell.alignment = STYLES["header_alignment"]

    ws.freeze_panes = "A2"


def _auto_size_columns(ws: Worksheet, max_width: int = 80) -> None:
    """Auto-size columns based on content, capped at max_width."""
    for col_cells in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col_cells[0].column)  # type: ignore[union-attr]
        for cell in col_cells:
            if cell.value:  # type: ignore[union-attr]
                length = len(str(cell.value))  # type: ignore[union-attr]
                max_length = max(max_length, length)
        adjusted = min(max_length + 2, max_width)
        ws.column_dimensions[col_letter].width = adjusted


# ── Public API ────────────────────────────────────────────


def export_xlsx(
    data: dict[str, Any],
    run_id: str,
    reports_dir: Path,
) -> Path:
    """Export structured report data to an Excel workbook.

    Parameters
    ----------
    data
        Dict with keys: ``executive_summary``, ``params``, ``generated_at``,
        ``summaries``, ``trends``.
    run_id
        Unique identifier for the pipeline run.
    reports_dir
        Base directory for report outputs.

    Returns
    -------
    Path
        Absolute path to the written .xlsx file.

    Raises
    ------
    pydantic.ValidationError
        If the generated file fails validation.
    """
    out_dir = reports_dir / run_id
    out_dir.mkdir(parents=True, exist_ok=True)

    wb = Workbook()

    # Build sheets (first sheet is auto-created, rename it to Summary)
    summary_ws = wb.active
    assert summary_ws is not None
    _build_summary_sheet(summary_ws, data)

    stories_ws = wb.create_sheet()
    _build_stories_sheet(stories_ws, data.get("summaries", []))

    sources_ws = wb.create_sheet()
    _build_sources_sheet(sources_ws, data.get("summaries", []))

    trends_ws = wb.create_sheet()
    _build_trends_sheet(trends_ws, data.get("trends", []))

    file_path = (out_dir / "report.xlsx").resolve()
    wb.save(str(file_path))
    wb.close()

    # Validate output
    XlsxOutput(file_path=file_path)

    logger.info(
        "xlsx_exported",
        run_id=run_id,
        path=str(file_path),
        size_bytes=file_path.stat().st_size,
        sheet_count=4,
    )

    return file_path
